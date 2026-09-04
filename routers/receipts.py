import asyncio
import csv
from hashlib import sha256
import io
import logging
import mimetypes
import os
import re
import secrets
import tempfile
import uuid
import zipfile
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageOps, UnidentifiedImageError
import pillow_heif
pillow_heif.register_heif_opener()
from sqlalchemy import extract, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from database import get_db
from models import Invitation, Receipt, SecurityEvent, User
from services.auth import (
    authenticate_user,
    clear_auth_cookie,
    create_user_session,
    get_authenticated_user,
    hash_password,
    login_is_rate_limited,
    normalize_email,
    normalize_username,
    record_security_event,
    revoke_all_user_sessions,
    revoke_current_session,
    set_auth_cookie,
    token_hash,
    utcnow,
    validate_password,
)
from services.mock_analysis import MOCK_RESULT, MOCK_RESULTS_BY_FILENAME, analyze_receipt

JST = timezone(timedelta(hours=9))

router = APIRouter()
templates = Jinja2Templates(directory="templates")
PORTFOLIO_DEMO_MODE = os.getenv("PORTFOLIO_DEMO_MODE", "false").lower() == "true" and os.getenv("APP_ENV", "development").lower() != "test"
templates.env.globals["portfolio_demo_mode"] = PORTFOLIO_DEMO_MODE


def _to_jst(dt) -> str:
    if dt is None:
        return ""
    return dt.replace(tzinfo=timezone.utc).astimezone(JST).strftime("%Y/%m/%d %H:%M")


templates.env.filters["jst"] = _to_jst

if PORTFOLIO_DEMO_MODE:
    # Vercel's deployed bundle is read-only; only /tmp is writable.
    demo_storage_root = Path(tempfile.gettempdir()) / "receipt-expense-ai"
    PRIVATE_UPLOAD_FOLDER = str(demo_storage_root / "uploads")
    LEGACY_UPLOAD_FOLDER = str(demo_storage_root / "legacy-uploads")
else:
    PRIVATE_UPLOAD_FOLDER = "storage/uploads"
    LEGACY_UPLOAD_FOLDER = "static/uploads"
MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))
MAX_FILES_PER_UPLOAD = int(os.getenv("MAX_FILES_PER_UPLOAD", "10"))
MAX_CONCURRENT_ANALYSES = max(1, int(os.getenv("MAX_CONCURRENT_ANALYSES", "5")))
ANALYSIS_TIMEOUT_SECONDS = float(os.getenv("ANALYSIS_TIMEOUT_SECONDS", "50"))
UPLOAD_RATE_LIMIT_WINDOW = 10 * 60
logger = logging.getLogger(__name__)
# JPEGとして保存できるフォーマット（iOSのMPO・HEIFなど含む）
JPEG_COMPATIBLE_FORMATS = {"JPEG", "MPO", "HEIF", "HEIC", "TIFF", "BMP", "GIF"}
# PNG/WEBPはそのまま保存
PNG_FORMATS = {"PNG"}
WEBP_FORMATS = {"WEBP"}
ALLOWED_IMAGE_FORMATS = JPEG_COMPATIBLE_FORMATS | PNG_FORMATS | WEBP_FORMATS
MAX_IMAGE_PIXELS = int(os.getenv("MAX_IMAGE_PIXELS", "25000000"))
INVITATION_VALID_HOURS = int(os.getenv("INVITATION_VALID_HOURS", str(7 * 24)))

PORTFOLIO_DEMO_SAMPLE_FILES = (
    ("portfolio-demo-grocery.png", "receipt-grocery.png"),
    ("portfolio-demo-cafe.png", "receipt-cafe.png"),
    ("portfolio-demo-stationery.png", "receipt-stationery.png"),
    ("portfolio-demo-restaurant.png", "receipt-restaurant.png"),
    ("portfolio-demo-household.png", "receipt-household.png"),
)

os.makedirs(PRIVATE_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(LEGACY_UPLOAD_FOLDER, exist_ok=True)


# ─────────────────────────── Security Helpers ───────────────────────────

def get_csrf_token(request: Request) -> str:
    token = request.session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        request.session["csrf_token"] = token
    return token


def validate_csrf(request: Request, csrf_token: str) -> None:
    session_token = request.session.get("csrf_token")
    if not csrf_token or not session_token or not secrets.compare_digest(csrf_token, session_token):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="不正なリクエストです。ページを再読み込みしてやり直してください。")


def render_template(request: Request, template_name: str, context: dict | None = None, status_code: int = 200):
    merged = dict(context or {})
    merged["csrf_token"] = get_csrf_token(request)
    return templates.TemplateResponse(request, template_name, merged, status_code=status_code)


def get_portfolio_demo_user(db: Session) -> User | None:
    _, username_normalized = normalize_username(os.getenv("APP_USERNAME", "demo"))
    return (
        db.query(User)
        .filter(User.username_normalized == username_normalized, User.is_active.is_(True))
        .first()
    )


def ensure_authenticated(request: Request, db: Session) -> User:
    user = get_portfolio_demo_user(db) if PORTFOLIO_DEMO_MODE else get_authenticated_user(db, request)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            detail="ログインが必要です。",
            headers={"Location": "/login"},
        )
    if user.must_change_password and request.url.path not in {"/account/password", "/logout"}:
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            detail="パスワードの変更が必要です。",
            headers={"Location": "/account/password"},
        )
    return user


def ensure_admin(request: Request, db: Session) -> User:
    if PORTFOLIO_DEMO_MODE:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="この画面は公開デモでは使用しません。")
    user = ensure_authenticated(request, db)
    if user.role != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="管理者権限が必要です。")
    return user


def enforce_user_rate_limit(db: Session, request: Request, user: User, event_type: str, limit: int) -> None:
    since = utcnow() - timedelta(seconds=UPLOAD_RATE_LIMIT_WINDOW)
    count = (
        db.query(func.count(SecurityEvent.id))
        .filter(
            SecurityEvent.event_type == event_type,
            SecurityEvent.user_id == user.id,
            SecurityEvent.created_at >= since,
        )
        .scalar()
        or 0
    )
    if count >= limit:
        raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail="操作回数が多すぎます。しばらく待ってから再試行してください。")
    record_security_event(db, event_type, request, user_id=user.id)
    db.commit()


# ─────────────────────────── Upload Helpers ───────────────────────────

def safe_filename(filename: str) -> str:
    filename = os.path.basename(filename)
    filename = re.sub(r"[^\w\s.\-]", "", filename)
    return filename.strip() or "upload"


def build_content_disposition(filename: str, fallback_filename: str) -> str:
    return f'attachment; filename="{fallback_filename}"; filename*=UTF-8\'\'{quote(filename)}'


def normalize_image_bytes(contents: bytes) -> tuple[bytes, str]:
    if not contents:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="空のファイルはアップロードできません。")
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail=f"ファイルサイズは {MAX_UPLOAD_BYTES // (1024 * 1024)}MB 以下にしてください。")

    try:
        with Image.open(io.BytesIO(contents)) as image:
            source_format = (image.format or "").upper()
            if source_format not in ALLOWED_IMAGE_FORMATS:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="JPEG・PNG・WEBP・HEIF・HEIC形式の画像を選択してください。",
                )
            if image.width <= 0 or image.height <= 0 or image.width * image.height > MAX_IMAGE_PIXELS:
                raise HTTPException(
                    status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail="画像の縦横サイズが大きすぎます。",
                )

            # Pillowで開けた画像はすべて受け入れ、形式に応じてJPEG/PNG/WEBPに変換
            if source_format in PNG_FORMATS:
                save_format, extension = "PNG", ".png"
            elif source_format in WEBP_FORMATS:
                save_format, extension = "WEBP", ".webp"
            else:
                # JPEG互換（MPO, HEIF, BMP, TIFFなどiOS/Android各種）→ JPEG
                save_format, extension = "JPEG", ".jpg"

            normalized = ImageOps.exif_transpose(image)

            if save_format == "JPEG" and normalized.mode not in ("RGB", "L"):
                normalized = normalized.convert("RGB")
            elif save_format in {"PNG", "WEBP"} and normalized.mode == "P":
                normalized = normalized.convert("RGBA")

            # 長辺1600px以内にリサイズ（解析処理高速化 & DB節約）
            if max(normalized.width, normalized.height) > 1600:
                normalized.thumbnail((1600, 1600), Image.LANCZOS)

            output = io.BytesIO()
            if save_format == "JPEG":
                normalized.save(output, format="JPEG", quality=85, optimize=True)
            elif save_format == "PNG":
                normalized.save(output, format="PNG", optimize=True)
            elif save_format == "WEBP":
                normalized.save(output, format="WEBP", quality=85, method=6)

            return output.getvalue(), extension
    except HTTPException:
        raise
    except (UnidentifiedImageError, Image.DecompressionBombError, OSError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="画像ファイルとして認識できませんでした。別の画像をお試しください。") from exc


def ensure_portfolio_demo_receipts(db: Session) -> None:
    """Seed the local public demo with fictional receipts for date filtering."""
    if not PORTFOLIO_DEMO_MODE:
        return

    current_user = get_portfolio_demo_user(db)
    if not current_user:
        return

    demo_asset_root = Path(__file__).resolve().parents[1] / "assets" / "demo" / "receipts"
    changed = False

    for filename, asset_filename in PORTFOLIO_DEMO_SAMPLE_FILES:
        result = MOCK_RESULTS_BY_FILENAME.get(filename, MOCK_RESULT)
        receipt_date = extract_receipt_date(result, date.today().year)
        if not receipt_date:
            logger.warning("Portfolio demo receipt has no date: %s", filename)
            continue
        demo_uploaded_at = datetime(receipt_date.year, receipt_date.month, receipt_date.day, 9, 0, 0)

        existing = (
            db.query(Receipt)
            .filter(Receipt.owner_id == current_user.id, Receipt.filename == filename)
            .order_by(Receipt.id.asc())
            .first()
        )
        if existing:
            # Repair rows created before sample-specific Mock results were added,
            # while leaving any later user edits untouched.
            if existing.result == MOCK_RESULT:
                existing.result = result
                existing.receipt_date = receipt_date
                existing.uploaded_at = demo_uploaded_at
                changed = True
            elif existing.result == result:
                if existing.receipt_date != receipt_date:
                    existing.receipt_date = receipt_date
                    changed = True
                if existing.uploaded_at != demo_uploaded_at:
                    existing.uploaded_at = demo_uploaded_at
                    changed = True
            continue

        asset_path = demo_asset_root / asset_filename
        try:
            normalized_bytes, extension = normalize_image_bytes(asset_path.read_bytes())
        except (OSError, HTTPException) as exc:
            logger.warning("Portfolio demo receipt asset is unavailable: %s (%s)", asset_path, exc)
            continue

        image_hash = sha256(normalized_bytes).hexdigest()
        duplicate = (
            db.query(Receipt)
            .filter(Receipt.owner_id == current_user.id, Receipt.image_hash == image_hash)
            .first()
        )
        if duplicate:
            continue

        db.add(
            Receipt(
                owner_id=current_user.id,
                filename=filename,
                stored_filename=None,
                image_hash=image_hash,
                result=result,
                receipt_date=receipt_date,
                uploaded_at=demo_uploaded_at,
                image_data=normalized_bytes,
                image_content_type=mimetypes.guess_type(asset_path.name)[0] or f"image/{extension.lstrip('.')}",
            )
        )
        changed = True

    if changed:
        db.commit()


def make_stored_filename(extension: str) -> str:
    return f"{uuid.uuid4().hex}{extension}"


def resolve_receipt_file_path(receipt: Receipt | dict) -> str | None:
    stored_filename = receipt["stored_filename"] if isinstance(receipt, dict) else receipt.stored_filename
    original_filename = receipt["filename"] if isinstance(receipt, dict) else receipt.filename

    candidates: list[str] = []
    if stored_filename:
        candidates.append(os.path.join(PRIVATE_UPLOAD_FOLDER, stored_filename))
    if original_filename:
        candidates.append(os.path.join(PRIVATE_UPLOAD_FOLDER, original_filename))
        candidates.append(os.path.join(LEGACY_UPLOAD_FOLDER, original_filename))

    for candidate in candidates:
        if os.path.isfile(candidate):
            return candidate
    return None


def get_receipt_or_404(db: Session, receipt_id: int, owner_id: int) -> Receipt:
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id, Receipt.owner_id == owner_id).first()
    if not receipt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="レシートが見つかりません。")
    return receipt


# ─────────────────────────── Receipt Helpers ───────────────────────────

def get_receipt_reference_date(receipt_data: dict) -> date:
    return receipt_data["receipt_date"] or receipt_data["uploaded_at"].date()


def ensure_unique_path(path: str, used_paths: set[str]) -> str:
    base, ext = os.path.splitext(path)
    candidate = path
    counter = 2
    while candidate in used_paths:
        candidate = f"{base}_{counter}{ext}"
        counter += 1
    used_paths.add(candidate)
    return candidate


def extract_receipt_date(result: str, upload_year: int) -> date | None:
    match = re.search(r"日付：(\d{4})年(\d{1,2})月(\d{1,2})日", result)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            pass
    match = re.search(r"日付：(\d{1,2})月(\d{1,2})日", result)
    if match:
        month, day = int(match.group(1)), int(match.group(2))
        year = upload_year - 1 if month > date.today().month else upload_year
        try:
            return date(year, month, day)
        except ValueError:
            pass
    return None


def parse_receipt_fields(result: str) -> dict:
    def get(pattern):
        match = re.search(pattern, result)
        return match.group(1).strip() if match else "－"

    return {
        "date_str": get(r"日付：(.+)"),
        "store": get(r"お店、会社名：(.+)"),
        "category": get(r"勘定科目：(.+)"),
        "total": get(r"合計金額：(.+)"),
    }


def extract_edit_details(result: str) -> str:
    """Return the OCR body without the fields edited in the structured form."""
    detail_lines = []
    for line in (result or "").splitlines():
        if re.match(r"^(日付|お店、会社名|勘定科目|合計金額)：", line.strip()):
            continue
        detail_lines.append(line)
    return "\n".join(detail_lines).strip()


def compose_receipt_result(date_str: str, store: str, category: str, total: str, details: str) -> str:
    """Keep the existing OCR text format while allowing field-by-field editing."""
    clean = lambda value: re.sub(r"\s+", " ", (value or "").replace("\r", "").replace("\n", " ")).strip() or "－"
    detail_lines = []
    for line in (details or "").splitlines():
        if line.strip().startswith("合計金額："):
            continue
        detail_lines.append(line.rstrip())
    body = "\n".join(detail_lines).strip()
    fields = [
        f"日付：{clean(date_str)}",
        f"お店、会社名：{clean(store)}",
        f"勘定科目：{clean(category)}",
    ]
    if body:
        fields.append(body)
    fields.append(f"合計金額：{clean(total)}")
    return "\n".join(fields)


def parse_amount(total_str: str) -> int | None:
    match = re.search(r"[\d,]+", total_str)
    if match:
        try:
            return int(match.group().replace(",", ""))
        except ValueError:
            pass
    return None


def safe_spreadsheet_cell(value) -> str:
    text_value = str(value or "")
    check_value = text_value.lstrip("\t\r\n ")
    if check_value.startswith(("=", "+", "-", "@")):
        return "'" + text_value
    return text_value


def build_receipt_data(receipts_raw: list[Receipt]) -> list[dict]:
    result: list[dict] = []
    for receipt in receipts_raw:
        fields = parse_receipt_fields(receipt.result)
        reference_date = receipt.receipt_date or receipt.uploaded_at.date()
        result.append({
            "id": receipt.id,
            "filename": receipt.filename,
            "stored_filename": receipt.stored_filename,
            "uploaded_at": receipt.uploaded_at,
            "receipt_date": receipt.receipt_date,
            "result": receipt.result,
            "month_group": f"{reference_date.year}年{reference_date.month}月",
            "is_expense": receipt.is_expense if receipt.is_expense is not None else True,
            "total_int": parse_amount(fields["total"]),
            **fields,
        })
    return result


def compute_totals(receipts_data: list[dict]) -> tuple[str, dict]:
    year_sum = 0
    month_sums: dict[str, int] = {}
    for receipt in receipts_data:
        if not receipt["is_expense"]:
            continue
        amount = receipt["total_int"]
        if amount is not None:
            year_sum += amount
            month_sums[receipt["month_group"]] = month_sums.get(receipt["month_group"], 0) + amount

    def format_currency(value: int) -> str:
        return f"¥{value:,}"

    return format_currency(year_sum), {k: format_currency(v) for k, v in month_sums.items()}


def get_available_years(db: Session, owner_id: int) -> list[int]:
    rows = (
        db.query(func.extract("year", Receipt.receipt_date).label("y"))
        .filter(Receipt.owner_id == owner_id, Receipt.receipt_date.isnot(None))
        .distinct()
        .all()
    )
    return sorted([int(row.y) for row in rows if row.y], reverse=True)


# ─────────────────────────── Auth ───────────────────────────

@router.get("/login")
async def login_page(request: Request, db: Session = Depends(get_db), registered: int = Query(default=0)):
    if PORTFOLIO_DEMO_MODE:
        return RedirectResponse(url="/", status_code=303)
    if get_authenticated_user(db, request):
        return RedirectResponse(url="/receipts", status_code=303)
    return render_template(request, "login.html", {"registered": bool(registered)})


@router.post("/login")
async def login(
    request: Request,
    identifier: str = Form(...),
    password: str = Form(...),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    validate_csrf(request, csrf_token)
    identifier = (identifier or "").strip()[:320]
    if login_is_rate_limited(db, request, identifier):
        return render_template(
            request,
            "login.html",
            {"error": "ログイン試行回数が多すぎます。10分ほど待ってから再試行してください。"},
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
        )

    user = authenticate_user(db, identifier, password)
    if not user:
        record_security_event(db, "login_failure", request, identifier=identifier)
        db.commit()
        return render_template(
            request,
            "login.html",
            {"error": "ログイン情報が正しくありません。"},
            status_code=status.HTTP_401_UNAUTHORIZED,
        )

    request.session.clear()
    request.session["csrf_token"] = secrets.token_urlsafe(32)
    raw_token = create_user_session(db, user, request)
    user.last_login_at = utcnow()
    record_security_event(db, "login_success", request, user_id=user.id)
    db.commit()
    destination = "/account/password" if user.must_change_password else "/receipts"
    response = RedirectResponse(url=destination, status_code=303)
    set_auth_cookie(response, raw_token)
    return response


@router.post("/logout")
async def logout(request: Request, csrf_token: str = Form(default=""), db: Session = Depends(get_db)):
    if PORTFOLIO_DEMO_MODE:
        return RedirectResponse(url="/", status_code=303)
    user = ensure_authenticated(request, db)
    validate_csrf(request, csrf_token)
    revoke_current_session(db, request)
    record_security_event(db, "logout", request, user_id=user.id)
    db.commit()
    request.session.clear()
    response = RedirectResponse(url="/login", status_code=303)
    clear_auth_cookie(response)
    return response


@router.get("/account/password")
async def change_password_page(request: Request, db: Session = Depends(get_db)):
    if PORTFOLIO_DEMO_MODE:
        return RedirectResponse(url="/", status_code=303)
    user = ensure_authenticated(request, db)
    return render_template(request, "change_password.html", {"current_user": user})


@router.post("/account/password")
async def change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    new_password_confirm: str = Form(...),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    user = ensure_authenticated(request, db)
    validate_csrf(request, csrf_token)
    if not authenticate_user(db, user.username, current_password):
        return render_template(
            request,
            "change_password.html",
            {"current_user": user, "error": "現在のパスワードが正しくありません。"},
            status_code=400,
        )
    if new_password != new_password_confirm:
        return render_template(
            request,
            "change_password.html",
            {"current_user": user, "error": "新しいパスワードが一致しません。"},
            status_code=400,
        )
    try:
        user.password_hash = hash_password(new_password)
    except ValueError as exc:
        return render_template(
            request,
            "change_password.html",
            {"current_user": user, "error": str(exc)},
            status_code=400,
        )
    user.must_change_password = False
    revoke_all_user_sessions(db, user.id)
    raw_token = create_user_session(db, user, request)
    record_security_event(db, "password_changed", request, user_id=user.id)
    db.commit()
    response = RedirectResponse(url="/receipts?password_updated=1", status_code=303)
    set_auth_cookie(response, raw_token)
    return response


def get_valid_invitation(db: Session, raw_token: str) -> Invitation | None:
    if not raw_token or len(raw_token) > 200:
        return None
    return (
        db.query(Invitation)
        .filter(
            Invitation.token_hash == token_hash(raw_token),
            Invitation.used_at.is_(None),
            Invitation.expires_at > utcnow(),
        )
        .first()
    )


@router.get("/register")
async def register_page(request: Request, token: str = Query(default=""), db: Session = Depends(get_db)):
    if PORTFOLIO_DEMO_MODE:
        return RedirectResponse(url="/", status_code=303)
    invitation = get_valid_invitation(db, token)
    if not invitation:
        return render_template(
            request,
            "register.html",
            {"invalid_invitation": True},
            status_code=status.HTTP_400_BAD_REQUEST,
        )
    return render_template(request, "register.html", {"invitation": invitation, "invite_token": token})


@router.post("/register")
async def register(
    request: Request,
    invite_token: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    password_confirm: str = Form(...),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    validate_csrf(request, csrf_token)
    invitation = get_valid_invitation(db, invite_token)
    if not invitation:
        return render_template(
            request,
            "register.html",
            {"invalid_invitation": True},
            status_code=status.HTTP_400_BAD_REQUEST,
        )
    try:
        username_display, username_normalized = normalize_username(username)
        email_display, email_normalized = normalize_email(invitation.email)
        validate_password(password)
        if password != password_confirm:
            raise ValueError("パスワードが一致しません。")
    except ValueError as exc:
        return render_template(
            request,
            "register.html",
            {"invitation": invitation, "invite_token": invite_token, "error": str(exc), "entered_username": username},
            status_code=400,
        )

    if db.query(User.id).filter(User.username_normalized == username_normalized).first():
        error = "このユーザー名は使用できません。"
    elif db.query(User.id).filter(User.email_normalized == email_normalized).first():
        error = "このメールアドレスは既に登録されています。"
    else:
        error = None
    if error:
        return render_template(
            request,
            "register.html",
            {"invitation": invitation, "invite_token": invite_token, "error": error, "entered_username": username},
            status_code=400,
        )

    user = User(
        username=username_display,
        username_normalized=username_normalized,
        email=email_display,
        email_normalized=email_normalized,
        password_hash=hash_password(password),
        email_verified_at=utcnow(),
    )
    db.add(user)
    db.flush()
    invitation.used_at = utcnow()
    record_security_event(db, "registration_completed", request, user_id=user.id)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return render_template(
            request,
            "register.html",
            {
                "invitation": invitation,
                "invite_token": invite_token,
                "error": "登録情報が既に使用されています。管理者へ確認してください。",
                "entered_username": username,
            },
            status_code=409,
        )
    return RedirectResponse(url="/login?registered=1", status_code=303)


@router.get("/admin/users")
async def admin_users(request: Request, db: Session = Depends(get_db)):
    current_user = ensure_admin(request, db)
    users = db.query(User).order_by(User.created_at.asc()).all()
    invitations = (
        db.query(Invitation)
        .filter(Invitation.used_at.is_(None), Invitation.expires_at > utcnow())
        .order_by(Invitation.created_at.desc())
        .all()
    )
    return render_template(
        request,
        "admin_users.html",
        {"current_user": current_user, "users": users, "invitations": invitations},
    )


@router.post("/admin/invitations")
async def create_invitation(
    request: Request,
    email: str = Form(...),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    current_user = ensure_admin(request, db)
    validate_csrf(request, csrf_token)
    try:
        email_display, email_normalized = normalize_email(email)
    except ValueError as exc:
        return render_template(
            request,
            "admin_users.html",
            {
                "current_user": current_user,
                "users": db.query(User).order_by(User.created_at.asc()).all(),
                "invitations": db.query(Invitation).filter(Invitation.used_at.is_(None), Invitation.expires_at > utcnow()).all(),
                "error": str(exc),
            },
            status_code=400,
        )
    if db.query(User.id).filter(User.email_normalized == email_normalized).first():
        error = "このメールアドレスは既に登録されています。"
        return render_template(
            request,
            "admin_users.html",
            {
                "current_user": current_user,
                "users": db.query(User).order_by(User.created_at.asc()).all(),
                "invitations": db.query(Invitation).filter(Invitation.used_at.is_(None), Invitation.expires_at > utcnow()).all(),
                "error": error,
            },
            status_code=400,
        )
    raw_token = secrets.token_urlsafe(40)
    invitation = Invitation(
        email=email_display,
        email_normalized=email_normalized,
        token_hash=token_hash(raw_token),
        created_by_id=current_user.id,
        expires_at=utcnow() + timedelta(hours=INVITATION_VALID_HOURS),
    )
    db.add(invitation)
    record_security_event(db, "invitation_created", request, user_id=current_user.id)
    db.commit()
    return render_template(
        request,
        "admin_users.html",
        {
            "current_user": current_user,
            "users": db.query(User).order_by(User.created_at.asc()).all(),
            "invitations": db.query(Invitation).filter(Invitation.used_at.is_(None), Invitation.expires_at > utcnow()).all(),
            "invite_url": f"/register?token={quote(raw_token)}",
            "invite_email": email_display,
        },
    )


@router.post("/admin/users/{user_id}/toggle-active")
async def toggle_user_active(
    user_id: int,
    request: Request,
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    current_user = ensure_admin(request, db)
    validate_csrf(request, csrf_token)
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="自分自身を停止することはできません。")
    target = db.query(User).filter(User.id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="ユーザーが見つかりません。")
    target.is_active = not target.is_active
    if not target.is_active:
        revoke_all_user_sessions(db, target.id)
    record_security_event(db, "user_status_changed", request, user_id=current_user.id)
    db.commit()
    return RedirectResponse(url="/admin/users", status_code=303)


# ─────────────────────────── Pages ───────────────────────────

@router.get("/")
async def index(request: Request, db: Session = Depends(get_db)):
    current_user = ensure_authenticated(request, db)
    if current_user.must_change_password:
        return RedirectResponse(url="/account/password", status_code=303)
    return render_template(request, "index.html", {"current_user": current_user})


@router.get("/upload")
async def upload_page(request: Request, db: Session = Depends(get_db)):
    current_user = ensure_authenticated(request, db)
    if current_user.must_change_password:
        return RedirectResponse(url="/account/password", status_code=303)
    return render_template(
        request,
        "upload.html",
        {
            "current_user": current_user,
            "max_files_per_upload": MAX_FILES_PER_UPLOAD,
            "max_upload_mb": MAX_UPLOAD_BYTES // (1024 * 1024),
        },
    )


@router.post("/upload")
async def upload(
    request: Request,
    files: list[UploadFile] = File(...),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    current_user = ensure_authenticated(request, db)
    if current_user.must_change_password:
        return RedirectResponse(url="/account/password", status_code=303)
    validate_csrf(request, csrf_token)
    enforce_user_rate_limit(db, request, current_user, "upload_attempt", limit=20)

    if not files:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ファイルを選択してください。")
    if len(files) > MAX_FILES_PER_UPLOAD:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"同時アップロードは {MAX_FILES_PER_UPLOAD} 件までです。")

    errors = []
    duplicate_messages = []
    prepared_uploads = []
    pending_hashes = set()
    for file in files:
        original_filename = safe_filename(file.filename or "upload")
        contents = await file.read()
        try:
            normalized_bytes, extension = normalize_image_bytes(contents)
        except HTTPException as exc:
            errors.append(f"{original_filename}: {exc.detail}")
            continue

        stored_filename = make_stored_filename(extension)
        mime = mimetypes.guess_type(f"x{extension}")[0] or "application/octet-stream"
        image_hash = sha256(normalized_bytes).hexdigest()
        existing = (
            db.query(Receipt.id, Receipt.filename)
            .filter(Receipt.owner_id == current_user.id, Receipt.image_hash == image_hash)
            .first()
        )
        if existing or image_hash in pending_hashes:
            duplicate_messages.append(f"{original_filename} は既に登録済みのためスキップしました")
            continue
        pending_hashes.add(image_hash)
        prepared_uploads.append((original_filename, stored_filename, mime, normalized_bytes, image_hash))

    if not prepared_uploads:
        return render_template(
            request, "upload.html",
            {
                "error": "、".join(errors + duplicate_messages),
                "current_user": current_user,
                "max_files_per_upload": MAX_FILES_PER_UPLOAD,
                "max_upload_mb": MAX_UPLOAD_BYTES // (1024 * 1024),
            },
            status_code=400,
        )

    semaphore = asyncio.Semaphore(min(MAX_CONCURRENT_ANALYSES, len(prepared_uploads)))

    async def analyze_upload(upload_data):
        original_filename, stored_filename, mime, normalized_bytes, image_hash = upload_data
        try:
            async with semaphore:
                result = await asyncio.wait_for(
                    analyze_receipt(normalized_bytes, original_filename),
                    timeout=ANALYSIS_TIMEOUT_SECONDS,
                )
        except Exception:
            logger.exception("Receipt analysis failed: %s", original_filename)
            result = "AI解析に失敗しました。時間を置いて再度お試しください。"

        return original_filename, stored_filename, mime, normalized_bytes, image_hash, result

    analyzed_uploads = await asyncio.gather(
        *(analyze_upload(upload_data) for upload_data in prepared_uploads)
    )

    for original_filename, stored_filename, mime, normalized_bytes, image_hash, result in analyzed_uploads:
        receipt_date = extract_receipt_date(result, date.today().year)
        db.add(
            Receipt(
                owner_id=current_user.id,
                filename=original_filename,
                stored_filename=stored_filename,
                image_hash=image_hash,
                result=result,
                receipt_date=receipt_date,
                image_data=normalized_bytes,
                image_content_type=mime,
            )
        )

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        return render_template(
            request,
            "upload.html",
            {
                "current_user": current_user,
                "error": "同じ画像が同時に登録されました。重複分を除いて再度お試しください。",
                "max_files_per_upload": MAX_FILES_PER_UPLOAD,
                "max_upload_mb": MAX_UPLOAD_BYTES // (1024 * 1024),
            },
            status_code=409,
        )
    redirect_url = "/receipts"
    if duplicate_messages:
        redirect_url += f"?duplicate_count={len(duplicate_messages)}"
    return RedirectResponse(url=redirect_url, status_code=303)


@router.get("/receipts")
async def receipts_view(
    request: Request,
    db: Session = Depends(get_db),
    year: int = Query(default=None),
    month: int = Query(default=None),
    sort: str = Query(default="uploaded_at"),
    updated: int = Query(default=0),
    duplicate_count: int = Query(default=0),
    password_updated: int = Query(default=0),
):
    current_user = ensure_authenticated(request, db)
    if current_user.must_change_password:
        return RedirectResponse(url="/account/password", status_code=303)
    ensure_portfolio_demo_receipts(db)
    query = db.query(Receipt).filter(Receipt.owner_id == current_user.id)
    if year:
        query = query.filter(extract("year", Receipt.receipt_date) == year)
    if month:
        query = query.filter(extract("month", Receipt.receipt_date) == month)
    if sort == "receipt_date":
        query = query.order_by(Receipt.receipt_date.desc().nullslast())
    else:
        query = query.order_by(Receipt.uploaded_at.desc())

    receipts_data = build_receipt_data(query.all())
    year_total, month_totals = compute_totals(receipts_data)

    return render_template(
        request,
        "receipts.html",
        {
            "receipts": receipts_data,
            "current_user": current_user,
            "available_years": get_available_years(db, current_user.id),
            "selected_year": year,
            "selected_month": month,
            "sort": sort,
            "year_total": year_total,
            "month_totals": month_totals,
            "updated": bool(updated),
            "duplicate_count": duplicate_count,
            "password_updated": bool(password_updated),
        },
    )


@router.get("/receipts/{receipt_id}/image")
async def receipt_image(request: Request, receipt_id: int, db: Session = Depends(get_db)):
    current_user = ensure_authenticated(request, db)
    receipt = get_receipt_or_404(db, receipt_id, current_user.id)

    if receipt.image_data:
        from fastapi.responses import Response as _Response
        media_type = receipt.image_content_type or "application/octet-stream"
        return _Response(content=receipt.image_data, media_type=media_type)

    file_path = resolve_receipt_file_path(receipt)
    if not file_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="画像ファイルが見つかりません。")
    media_type = mimetypes.guess_type(file_path)[0] or "application/octet-stream"
    return FileResponse(file_path, media_type=media_type)


@router.post("/receipts/{receipt_id}/toggle-expense")
async def toggle_expense(
    receipt_id: int,
    request: Request,
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    current_user = ensure_authenticated(request, db)
    validate_csrf(request, csrf_token)
    receipt = get_receipt_or_404(db, receipt_id, current_user.id)
    receipt.is_expense = not (receipt.is_expense if receipt.is_expense is not None else True)
    db.commit()
    return RedirectResponse(url="/receipts", status_code=303)


@router.get("/receipts/export")
async def export_receipts(
    request: Request,
    db: Session = Depends(get_db),
    year: int = Query(default=None),
    month: int = Query(default=None),
):
    current_user = ensure_authenticated(request, db)
    query = db.query(Receipt).filter(Receipt.owner_id == current_user.id).order_by(Receipt.receipt_date.desc().nullslast())
    if year:
        query = query.filter(extract("year", Receipt.receipt_date) == year)
    if month:
        query = query.filter(extract("month", Receipt.receipt_date) == month)
    receipts_data = build_receipt_data(query.all())

    headers_row = [
        "レシート日付",
        "店舗名",
        "勘定科目",
        "合計金額",
        "経費対象",
        "ファイル名",
        "画像相対パス",
        "アップロード日時",
    ]

    receipt_ids = [receipt["id"] for receipt in receipts_data]
    receipt_objects = {
        receipt.id: receipt
        for receipt in db.query(Receipt)
        .filter(Receipt.owner_id == current_user.id, Receipt.id.in_(receipt_ids))
        .all()
    } if receipt_ids else {}

    used_image_paths: set[str] = set()
    export_rows = []
    for receipt in receipts_data:
        reference_date = get_receipt_reference_date(receipt)
        db_receipt = receipt_objects.get(receipt["id"])
        has_image = bool(
            (db_receipt and db_receipt.image_data)
            or resolve_receipt_file_path(receipt)
        )
        image_zip_path = None
        image_relative_path = "－"

        if has_image:
            image_zip_path = ensure_unique_path(
                f"images/{reference_date.year}/{reference_date.month:02d}/{safe_filename(receipt['filename'])}",
                used_image_paths,
            )
            image_relative_path = image_zip_path.replace("/", "\\")

        export_rows.append({
            **receipt,
            "image_zip_path": image_zip_path,
            "image_relative_path": image_relative_path,
            "_db_receipt": db_receipt,
        })

    def make_row(receipt: dict) -> list[str]:
        return [
            safe_spreadsheet_cell(value)
            for value in (
                receipt["date_str"],
                receipt["store"],
                receipt["category"],
                receipt["total"],
                "○" if receipt["is_expense"] else "×",
                receipt["filename"],
                receipt["image_relative_path"],
                _to_jst(receipt["uploaded_at"]),
            )
        ]

    label = f"_{year}年" if year else "_全期間"
    fallback_label = f"_{year}" if year else "_all"
    if month:
        label += f"{month}月"
        fallback_label += f"_{month:02d}"

    package_root = f"receipts{fallback_label}"
    xlsx_name = f"receipts{fallback_label}.xlsx"
    csv_name = f"receipts{fallback_label}.csv"
    content_disposition = build_content_disposition(
        filename=f"receipts{label}.zip",
        fallback_filename=f"{package_root}.zip",
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "レシート一覧"
    worksheet.append(headers_row)

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    image_path_col_index = headers_row.index("画像相対パス") + 1
    for receipt in export_rows:
        worksheet.append(make_row(receipt))
        if receipt["image_zip_path"]:
            cell = worksheet.cell(row=worksheet.max_row, column=image_path_col_index)
            cell.hyperlink = receipt["image_relative_path"]
            cell.font = Font(color="0563C1", underline="single")

    for col in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        worksheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    excel_output = io.BytesIO()
    workbook.save(excel_output)
    excel_output.seek(0)

    csv_output = io.StringIO()
    writer = csv.writer(csv_output)
    writer.writerow(headers_row)
    for receipt in export_rows:
        writer.writerow(make_row(receipt))

    zip_output = io.BytesIO()
    with zipfile.ZipFile(zip_output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(f"{package_root}/{xlsx_name}", excel_output.getvalue())
        archive.writestr(f"{package_root}/{csv_name}", csv_output.getvalue().encode("utf-8-sig"))
        for receipt in export_rows:
            if not receipt["image_zip_path"]:
                continue
            image_bytes: bytes | None = None
            db_receipt = receipt.get("_db_receipt")
            if db_receipt and db_receipt.image_data:
                image_bytes = db_receipt.image_data
            if image_bytes is None:
                local_path = resolve_receipt_file_path(receipt)
                if local_path:
                    with open(local_path, "rb") as img_f:
                        image_bytes = img_f.read()
            if image_bytes:
                archive.writestr(f"{package_root}/{receipt['image_zip_path']}", image_bytes)

    zip_output.seek(0)
    record_security_event(db, "receipts_exported", request, user_id=current_user.id)
    db.commit()
    return StreamingResponse(
        zip_output,
        media_type="application/zip",
        headers={"Content-Disposition": content_disposition},
    )


@router.get("/summary")
async def summary_page(
    request: Request,
    db: Session = Depends(get_db),
    year: int = Query(default=None),
):
    current_user = ensure_authenticated(request, db)
    query = db.query(Receipt).filter(Receipt.owner_id == current_user.id, Receipt.is_expense.is_(True))
    if year:
        query = query.filter(extract("year", Receipt.receipt_date) == year)
    receipts_data = build_receipt_data(query.all())

    category_totals: dict[str, dict] = {}
    month_totals: dict[str, int] = {}
    for receipt in receipts_data:
        category = receipt["category"]
        amount = receipt["total_int"] or 0
        if category not in category_totals:
            category_totals[category] = {"count": 0, "total": 0}
        category_totals[category]["count"] += 1
        category_totals[category]["total"] += amount

        month_group = receipt["month_group"]
        month_totals[month_group] = month_totals.get(month_group, 0) + amount

    grand_total = sum(value["total"] for value in category_totals.values())

    category_list = sorted(
        [
            {
                "category": key,
                "count": value["count"],
                "total": value["total"],
                "total_fmt": f"¥{value['total']:,}",
                "pct": round(value["total"] / grand_total * 100, 1) if grand_total else 0,
            }
            for key, value in category_totals.items()
        ],
        key=lambda item: item["total"],
        reverse=True,
    )

    month_list = sorted(
        [{"month": key, "total": value, "total_fmt": f"¥{value:,}"} for key, value in month_totals.items()],
        key=lambda item: item["month"],
    )

    return render_template(
        request,
        "summary.html",
        {
            "category_list": category_list,
            "current_user": current_user,
            "month_list": month_list,
            "grand_total": f"¥{grand_total:,}",
            "receipt_count": len(receipts_data),
            "available_years": get_available_years(db, current_user.id),
            "selected_year": year,
        },
    )


@router.post("/receipts/delete")
async def delete_receipts(
    request: Request,
    ids: list[int] = Form(default=[]),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    current_user = ensure_authenticated(request, db)
    validate_csrf(request, csrf_token)
    if ids:
        owned_receipts = db.query(Receipt).filter(Receipt.owner_id == current_user.id, Receipt.id.in_(ids)).all()
        for receipt in owned_receipts:
            file_path = resolve_receipt_file_path(receipt)
            if file_path and os.path.commonpath([os.path.abspath(file_path), os.path.abspath(PRIVATE_UPLOAD_FOLDER)]) == os.path.abspath(PRIVATE_UPLOAD_FOLDER):
                try:
                    os.remove(file_path)
                except FileNotFoundError:
                    pass
        owned_ids = [receipt.id for receipt in owned_receipts]
        if owned_ids:
            db.query(Receipt).filter(Receipt.owner_id == current_user.id, Receipt.id.in_(owned_ids)).delete(synchronize_session=False)
            record_security_event(db, "receipts_deleted", request, user_id=current_user.id)
        db.commit()
    return RedirectResponse(url="/receipts", status_code=303)


@router.get("/receipts/{receipt_id}/edit")
async def edit_receipt_page(receipt_id: int, request: Request, db: Session = Depends(get_db)):
    current_user = ensure_authenticated(request, db)
    receipt = get_receipt_or_404(db, receipt_id, current_user.id)
    return render_template(
        request,
        "edit_receipt.html",
        {
            "current_user": current_user,
            "receipt": receipt,
            "current_user": current_user,
            "edit_fields": parse_receipt_fields(receipt.result),
            "edit_details": extract_edit_details(receipt.result),
        },
    )


@router.post("/receipts/{receipt_id}/edit")
async def edit_receipt(
    receipt_id: int,
    request: Request,
    result: str = Form(default=""),
    date_str: str = Form(default=""),
    store: str = Form(default=""),
    category: str = Form(default=""),
    total: str = Form(default=""),
    details: str = Form(default=""),
    csrf_token: str = Form(default=""),
    db: Session = Depends(get_db),
):
    current_user = ensure_authenticated(request, db)
    validate_csrf(request, csrf_token)
    receipt = get_receipt_or_404(db, receipt_id, current_user.id)
    if any(value.strip() for value in (date_str, store, category, total, details)):
        result = compose_receipt_result(date_str, store, category, total, details)
    if not result.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="保存する内容を入力してください。")
    receipt.result = result
    receipt.receipt_date = extract_receipt_date(result, receipt.uploaded_at.year)
    db.commit()
    return RedirectResponse(url="/receipts?updated=1", status_code=303)
